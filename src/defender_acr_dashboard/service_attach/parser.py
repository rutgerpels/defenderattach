"""Parse the combined SL2/SL4 ACR export into a tidy long-format table.

The export has a two-row header: row 1 carries the fiscal-month group label
(merged across five measure columns), row 2 carries the real measure header.
Only the ``$ ACR`` measure of each month is retained. Streaming with
``iter_rows`` is required; random ``cell()`` access on a read-only workbook is
pathologically slow on this file.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd

from .mapping import DEFENDER_SL2, TOTAL_TOKEN

EXPORT_SHEET = "Export"
ACR_MEASURE = "$ ACR"
MONTH_GROUP_HEADER = "FiscalMonth"

LEVEL_CUSTOMER_TOTAL = "customer_total"
LEVEL_SERVICE_TOTAL = "service_total"
LEVEL_LEAF = "leaf"


@dataclass
class ReconciliationIssue:
    customer: str
    scope: str
    expected: float
    actual: float

    @property
    def abs_diff(self) -> float:
        return abs(self.expected - self.actual)

    @property
    def rel_diff(self) -> float:
        base = max(abs(self.expected), 1.0)
        return self.abs_diff / base


@dataclass
class ParsedData:
    """Tidy parse result.

    frame columns: customer, sl2, sl4, level, month, acr
    """

    frame: pd.DataFrame
    months: List[str]
    customers: List[str]
    reconciliation: List[ReconciliationIssue] = field(default_factory=list)
    source_name: str = ""
    row_count: int = 0

    @property
    def latest_month(self) -> Optional[str]:
        return self.months[-1] if self.months else None

    @property
    def reconciliation_ok(self) -> bool:
        return all(issue.rel_diff <= 0.01 for issue in self.reconciliation)


def _coerce_float(value: object) -> float:
    if value is None:
        return 0.0
    try:
        result = float(value)
    except (TypeError, ValueError):
        return 0.0
    if result != result:  # NaN
        return 0.0
    return result


def _month_columns(header_top: List[object], header_bottom: List[object]) -> List[Tuple[str, int]]:
    """Return ordered (month_label, column_index) pairs for ``$ ACR`` columns.

    The ``Total`` month group is excluded; it is a fiscal-year roll-up, not part
    of the monthly time series.
    """

    months: List[Tuple[str, int]] = []
    current_group: Optional[str] = None
    for idx, top in enumerate(header_top):
        if top is not None:
            current_group = str(top).strip()
        bottom = header_bottom[idx]
        if bottom is None:
            continue
        if str(bottom).strip() != ACR_MEASURE:
            continue
        if current_group in (None, MONTH_GROUP_HEADER, TOTAL_TOKEN):
            continue
        months.append((current_group, idx))
    return months


def _classify_level(sl2: Optional[str], sl4: Optional[str]) -> Tuple[str, str, str]:
    """Return (level, normalized_sl2, normalized_sl4)."""

    sl2_clean = (sl2 or "").strip()
    sl4_clean = (sl4 or "").strip()

    if sl2_clean == TOTAL_TOKEN and not sl4_clean:
        return LEVEL_CUSTOMER_TOTAL, TOTAL_TOKEN, ""
    if sl4_clean == TOTAL_TOKEN:
        return LEVEL_SERVICE_TOTAL, sl2_clean, TOTAL_TOKEN
    # Leaf rows where SL4 is blank still describe a real (single-leaf) service.
    if not sl4_clean:
        sl4_clean = sl2_clean
    return LEVEL_LEAF, sl2_clean, sl4_clean


def parse_sl2_sl4(path: Path | str) -> ParsedData:
    path = Path(path)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if EXPORT_SHEET in workbook.sheetnames:
            worksheet = workbook[EXPORT_SHEET]
        else:
            worksheet = workbook[workbook.sheetnames[0]]

        rows = worksheet.iter_rows(values_only=True)
        header_top = list(next(rows))
        header_bottom = list(next(rows))
        month_cols = _month_columns(header_top, header_bottom)
        months = [label for label, _ in month_cols]
        max_needed = (max((c for _, c in month_cols), default=2)) + 1

        records: List[dict] = []
        customers: List[str] = []
        seen_customers = set()
        row_count = 0

        for raw in rows:
            if raw is None:
                continue
            row = list(raw)
            if len(row) < 3:
                continue
            customer = (row[0] or "").strip() if isinstance(row[0], str) else row[0]
            if not customer:
                continue
            if len(row) < max_needed:
                row = row + [None] * (max_needed - len(row))

            level, sl2_clean, sl4_clean = _classify_level(row[1], row[2])
            row_count += 1

            if customer not in seen_customers:
                seen_customers.add(customer)
                customers.append(customer)

            for label, col in month_cols:
                records.append(
                    {
                        "customer": customer,
                        "sl2": sl2_clean,
                        "sl4": sl4_clean,
                        "level": level,
                        "month": label,
                        "acr": _coerce_float(row[col]),
                    }
                )
    finally:
        workbook.close()

    frame = pd.DataFrame.from_records(
        records, columns=["customer", "sl2", "sl4", "level", "month", "acr"]
    )
    reconciliation = _reconcile(frame, months)

    return ParsedData(
        frame=frame,
        months=months,
        customers=customers,
        reconciliation=reconciliation,
        source_name=path.name,
        row_count=row_count,
    )


def _reconcile(frame: pd.DataFrame, months: List[str]) -> List[ReconciliationIssue]:
    """Cross-check provided subtotal rows against independent leaf sums.

    Uses the latest month so the trust signal reflects current data.
    """

    issues: List[ReconciliationIssue] = []
    if frame.empty or not months:
        return issues

    latest = months[-1]
    snap = frame[frame["month"] == latest]

    for customer, cust_rows in snap.groupby("customer"):
        # Customer total row vs sum of per-service subtotals.
        total_row = cust_rows[cust_rows["level"] == LEVEL_CUSTOMER_TOTAL]["acr"].sum()
        service_totals = cust_rows[cust_rows["level"] == LEVEL_SERVICE_TOTAL]
        sum_service = service_totals["acr"].sum()
        if total_row:
            issues.append(
                ReconciliationIssue(
                    customer=customer,
                    scope="customer_total_vs_service_subtotals",
                    expected=float(total_row),
                    actual=float(sum_service),
                )
            )

        # Each service subtotal vs sum of its own leaves.
        leaves = cust_rows[cust_rows["level"] == LEVEL_LEAF]
        for sl2, sub_acr in service_totals.groupby("sl2")["acr"].sum().items():
            leaf_sum = leaves[leaves["sl2"] == sl2]["acr"].sum()
            if sub_acr and abs(sub_acr - leaf_sum) / max(abs(sub_acr), 1.0) > 0.01:
                issues.append(
                    ReconciliationIssue(
                        customer=customer,
                        scope=f"service_subtotal::{sl2}",
                        expected=float(sub_acr),
                        actual=float(leaf_sum),
                    )
                )

    return issues


def defender_plan_actuals(frame: pd.DataFrame) -> pd.DataFrame:
    """Leaf rows that represent actual Defender plan spend (one row per SL4)."""

    mask = (frame["sl2"] == DEFENDER_SL2) & (frame["level"] == LEVEL_LEAF)
    return frame[mask]
